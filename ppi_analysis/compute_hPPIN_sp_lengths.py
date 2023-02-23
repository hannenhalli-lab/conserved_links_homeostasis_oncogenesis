#!/usr/bin/python

import sys
import networkx as nx
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import math
import random
import statistics as stats
from scipy.stats import wilcoxon
import openpyxl
from pathlib import Path

wb_obj = openpyxl.load_workbook('hPPIN_network.xlsx') 
sheet = wb_obj.active

hPPIN = nx.Graph()
for row in sheet.iter_rows(2, sheet.max_row):
    node1 = row[0].value
    node2 = row[1].value
    hPPIN.add_edge(node1, node2, weight=1)
len(list(hPPIN.nodes))

# +
# create list
file2 = "common_nodes_hPPIN_STRING_PANCAN.txt"
combined_list = []
with open(file2) as IN:
    for line in IN:
        combined_list.append(line.rstrip().split("\t")[0])
        
removal_list = [node for node in hPPIN.nodes if node not in combined_list]
hPPIN.remove_nodes_from(removal_list)
# -

path_lengths = dict(nx.all_pairs_shortest_path_length(hPPIN))

with open("hPPIN_sp_lengths_total.txt", 'w') as f: 
    for key, value in path_lengths.items(): 
        f.write('%s\t%s\n' % (key, value))

# +
# Step 1
import pickle
 
# Step 2
with open('hPPIN_paths.dictionary', 'rb') as config_dictionary_file:
 
    # Step 3
    hPPIN_paths = pickle.load(config_dictionary_file)
 
# -


