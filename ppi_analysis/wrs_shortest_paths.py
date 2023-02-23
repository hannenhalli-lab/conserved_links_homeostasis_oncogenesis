#!/usr/bin/python

import sys
import networkx as nx
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import math
import random
import statistics as stats
from scipy.stats import wilcoxon
import openpyxl
from pathlib import Path

def list_contains(List1, List2): 
    check = []
  
    # add elements in list 1 that are in list2
    for m in List1: 
        if (m in List2) & (not (m in check)): 
            check.append(m)
            
    return check

# return only shortest paths involving both  nodes  of  interest
def get_sp_paths_length(G, noi):

    
    # list of elements in the graph
    node_list = list(G.nodes)
    
    # create dictionary with key path and value shortest path length
    path_dict = {}
    for nodex in noi:
        if (nodex in node_list):
            for nodey in noi:
                if ((nodex < nodey) & (nodey in node_list) & (nx.has_path(G, nodex, nodey))):
                    path = nodex + "-" + nodey
                    path_dict[path] = len(list(nx.dijkstra_path(G, nodex, nodey))) - 1
                    
                    #print the path
                    #print(list(nx.dijkstra_path(G, nodex, nodey)))
                    
    # Convert to a pandas dataframe
    spaths_df = pd.DataFrame.from_dict(path_dict, orient='index', columns = ['sp_length'])

    return spaths_df

# return only shortest paths involving both  nodes  of  interest
def get_random_paths_length(G, tissue_list):
    
    # list of elements in the graph
    node_list = list(G.nodes)
    gene_list = list_contains(tissue_list, node_list)
    n = 5
    
    # create dictionary with key path and value shortest path length
    path_dict = {}
    for i in range(n):
        not_unique = True
        while(not_unique):
            nodes = random.sample(gene_list, 2)
            nodex = nodes[0]
            nodey = nodes[1]
            
            if((nodex != nodey) & (nx.has_path(G, nodex, nodey))):
                not_unique  = False
                path = nodex + "-" + nodey
                path_dict[path] = len(list(nx.dijkstra_path(G, nodex, nodey))) - 1
                print(list(nx.dijkstra_path(G, nodex, nodey)))
               
                #print the path
                #print(list(nx.dijkstra_path(G, nodex, nodey)))
                    
    # Convert to a pandas dataframe
    spaths_df = pd.DataFrame.from_dict(path_dict, orient='index', columns = ['sp_length'])

    return spaths_df

# compare the distributions of the shorest path lengths for a  set and its matched random
def get_sp_network_dif(G, genes):
    
    # read in the genelist of interest
    file_genes = pd.read_csv(genes, header = None, names = ['genes'])
    
    # make sure that the genes are in the network
    node_list = list(G.nodes)
    gene_list = list_contains(file_genes['genes'], node_list)
    print(len(gene_list))
    
    # sample a random gene list to generate plot
    
    gene_path_lngths = get_sp_paths_length(G, gene_list)

    #labels = [name, 'Random']
    #plt.boxplot([list(gene_path_lngths['sp_length']), list(random_path_lngths['sp_length'])])
    #plt.title('Distribution of Shortest Path Lengths')
    #plt.xlabel('Gene Grouping')
    #plt.ylabel('Shortest Path Distribution')
    #plt.xticks([1,2], labels, rotation='vertical')

    #plt.show()
    
    return(gene_path_lngths)
# return only shortest paths involving both  nodes  of  interest in setA and setB
def get_sp_cancer_paths_length(G, file_genes_A, file_genes_B):
    
    # read in the genelist of interest
    noi_A = pd.read_csv(file_genes_A, header = None, names = ['genes'])
    noi_B = pd.read_csv(file_genes_B, header = None, names = ['genes'])
     # list of elements in the graph
    node_list = list(G.nodes)
    
    gene_list_A = list_contains(noi_A['genes'], node_list)
    gene_list_B = list_contains(noi_B['genes'], node_list)
    
    print(len(gene_list_A))
    print(len(gene_list_B))
    
    # create dictionary with key path and value shortest path length
    path_dict = {}
    for nodex in gene_list_A:
        if (nodex in node_list):
            for nodey in gene_list_B:
                if ((nodey in node_list) & (nx.has_path(G, nodex, nodey))):
                    path = nodex + "-" + nodey
                    path_dict[path] = len(list(nx.dijkstra_path(G, nodex, nodey))) - 1
                    
                    #print the path
                    #print(list(nx.dijkstra_path(G, nodex, nodey)))
                    
    # Convert to a pandas dataframe
    spaths_df = pd.DataFrame.from_dict(path_dict, orient='index', columns = ['sp_length'])

    return spaths_df

def plot_path_diff(files, names):
    
    # read in the genelist of interest
    boxplot_list= []
    for i in range(len(files)):
        current_file = pd.read_csv(files[i], sep="\t", index_col=0)['sp_length']
        print("N for " + names[i] + "  is " + str(len(current_file)) )
        print("Mean for " + names[i] + "  is " + str(stats.mean(current_file)) )
        print("Median for " + names[i] + "  is " + str(stats.median(current_file)))
        print("\n")
        
        boxplot_list.append(current_file)
    
    axes_list =  list(range(1, (len(names)+1)))
    
    labels = names
    plt.boxplot(boxplot_list)
    plt.title('Distribution of Shortest Path Lengths')
    plt.xlabel('Gene Grouping')
    plt.ylabel('Shortest Path Distribution')
    plt.xticks(axes_list, labels)

    plt.show()

wb_obj = openpyxl.load_workbook('hPPIN_network.xlsx') 
sheet = wb_obj.active

hPPIN = nx.Graph()
for row in sheet.iter_rows(2, sheet.max_row):
    node1 = row[0].value
    node2 = row[1].value
    hPPIN.add_edge(node1, node2, weight=1)
len(list(hPPIN.nodes))

# get shortest_path_lengths
matched_paths = get_sp_cancer_paths_length(hPPIN, sys.argv[1], sys.argv[2])

with open(sys.argv[3], 'a') as f:
    dfAsString = matched_paths.to_string(header=False, index=True)
    f.write(dfAsString)
