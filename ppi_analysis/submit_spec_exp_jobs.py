#!/usr/bin/python

import networkx as nx
import pandas as pd
import numpy as np
import random
import openpyxl
import math
import statistics as stats
import shlex
import time
import os

wb_obj = openpyxl.load_workbook('hPPIN_network.xlsx') 
sheet = wb_obj.active

hPPIN = nx.Graph()
for row in sheet.iter_rows(2, sheet.max_row):
    node1 = row[0].value
    node2 = row[1].value
    hPPIN.add_edge(node1, node2, weight=1)
len(list(hPPIN.nodes))

def genome_wrs_distance(genes, filename, filename_2,output):
    # split list into groups of 10
    sub_lists = [genes[i:i + 10] for i in range(0, len(genes), 10)]
    
    # write out matched control list
    for i in range(len(sub_lists)):
        file_in1 = "./genome_analysis/" + filename + str(i) + ".txt"
        file_in2 = "/data/timonaj/cancer_as_wound/ppi_analysis/" + filename_2 
        file_out = "./genome_analysis/" + output + str(i) + ".out"
        textfile = open(file_in1, 'w') 
        
        for element in sub_lists[i]:
            textfile.write(element + "\n")
        textfile.close()
            
        command = "printf \"#!/bin/bash\nmodule load python\npython /data/timonaj/cancer_as_wound/ppi_analysis/wrs_shortest_paths.py " + file_in1 + ' ' + file_in2 + ' ' + file_out + "\" > " + "./genome_analysis/jobscript" + str(i) + ".job \nsbatch --mem=1g --cpus-per-task=2 --gres=lscratch:2 --time 4:00:00 " + "./genome_analysis/jobscript" + str(i) + ".job \n"
        print(command)
        os.system(command)
        
        if i % 500 == 0 and i != 0:
            time.sleep(1200)

for file in os.listdir("/data/timonaj/cancer_as_wound/spec_exp_genes"):
    #genome_genes = [node for node in hPPIN.nodes()]
    print(file)
    output = "genome_" +  file[:-4]
    print(output)
    genome_genes = [node for node in hPPIN.nodes()]
    genome_wrs_distance(genome_genes, "genome_genes", file, output)
